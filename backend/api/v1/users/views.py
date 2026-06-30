import csv
import io
from pathlib import Path
from django.db import transaction
from datetime import datetime
from django.utils.crypto import get_random_string
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.viewsets import ModelViewSet
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.users.models import Election, Log, Organisation, PermissionRecord, User, Membership, ROLE_CHOICES
from apps.elections.models import Participant
from services.otp_service import issue_otp
from services.permission_service import assign_election_bulk_permissions_to_membership, assign_org_bulk_permissions_to_membership, revoke_org_bulk_permissions_from_membership, revoke_election_bulk_permissions_from_membership, get_all_permissions_for_membership, assign_org_default_permissions_to_membership, assign_election_default_permissions_to_membership
from .serializers import ElectionSerializer, LogSerializer, OrganisationSerializer, PermissionRecordSerializer, UserSerializer, MembershipSerializer, UserMembershipCreateSerializer
from .permissions import HasPermission, DenyAll
from services.membership_service import get_user_active_membership, get_user_active_organisation, switch_active_membership
from services.blockchain_service import deploy_election_contract
from services.voting_link_service import dispatch_voter_invites_for_election
from api.v1.elections.serializers import PositionSerializer, ParticipantSerializer, CandidateSerializer

ORG_ALLOWED_ROLE_CHOICES = {role for role, _ in ROLE_CHOICES}


def _normalize_column_name(value):
    return str(value or '').strip().lower().replace(' ', '_')


def _parse_csv_rows(uploaded_file):
    uploaded_file.seek(0)
    text = uploaded_file.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        if not row or not any(v not in (None, '') for v in row.values()):
            continue
        rows.append({_normalize_column_name(k): v for k, v in row.items()})
    return rows


def _parse_xlsx_rows(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('XLSX upload requires openpyxl to be installed.') from exc

    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    raw_rows = list(worksheet.iter_rows(values_only=True))
    if not raw_rows:
        return []

    headers = [_normalize_column_name(header) for header in raw_rows[0]]
    rows = []
    for row in raw_rows[1:]:
        if not row or not any(v not in (None, '') for v in row):
            continue
        parsed_row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            parsed_row[header] = row[idx] if idx < len(row) else None
        rows.append(parsed_row)
    return rows


def _parse_upload_rows(uploaded_file):
    extension = Path(uploaded_file.name).suffix.lower()
    if extension == '.csv':
        return _parse_csv_rows(uploaded_file)
    if extension in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
        return _parse_xlsx_rows(uploaded_file)
    raise ValueError('Unsupported file type. Please upload .csv or .xlsx.')


def _parse_bool(value, default=True):
    if value in (None, ''):
        return default
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on'}





class OrganisationViewSet(ModelViewSet):
    queryset = Organisation.objects.all()
    serializer_class = OrganisationSerializer
    permission_classes = [HasPermission]
    filterset_fields = ['name']

    ACTION_PERMISSION_MAP = {
        'create': 'org.manage',
        'update': 'org.manage',
        'partial_update': 'org.manage',
        'destroy': 'org.manage',
        'election_stats': 'org.analytics.view',
    }

    def get_permissions(self):
        if self.action == 'register_org':
            return [AllowAny()]
        if self.action in {'list', 'retrieve'}:
            return [IsAuthenticated()]
        return super().get_permissions()

    #in this case we want to fetch organisations a person is a member of
    def get_queryset(self):
        user = self.request.user
        
        return Organisation.objects.filter(memberships__user=user)

    @action(detail=False, methods=['post'], url_path='register-org')
    def register_org(self, request):
        organisation_name = str(request.data.get('organisation_name') or request.data.get('name') or '').strip()
        organisation_description = str(request.data.get('organisation_description') or request.data.get('description') or '').strip()
        email = str(request.data.get('email') or '').strip().lower()
        first_name = str(request.data.get('first_name') or '').strip()
        last_name = str(request.data.get('last_name') or '').strip()
        phone = str(request.data.get('phone') or '').strip()
        bio = str(request.data.get('bio') or '').strip()

        if not organisation_name:
            return Response({'detail': 'organisation_name is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if not email:
            return Response({'detail': 'email is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if User.objects.filter(email=email).exists():
            return Response({'detail': 'A user with this email already exists.'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            organisation = Organisation.objects.create(
                name=organisation_name,
                description=organisation_description or None,
            )
            user = User.objects.create_user(
                email=email,
                password=None,
                first_name=first_name,
                last_name=last_name,
                phone=phone or None,
                bio=bio or None,
                is_verified=False,
            )
            membership = Membership.objects.create(
                organisation=organisation,
                user=user,
                role='admin',
                currently_active=True,
                is_active=True,
            )
            assign_org_default_permissions_to_membership(membership.id, membership.role)
            transaction.on_commit(lambda: issue_otp(user))

        switch_active_membership(user.id, membership.id)
        return Response(
            {
                'detail': 'Organisation registered. A verification code has been sent to your email.',
                'organisation': OrganisationSerializer(organisation).data,
                'membership': MembershipSerializer(membership).data,
                'verification_sent': True,
            },
            status=status.HTTP_201_CREATED,
        )
    
    #an endpoint to get election statistics for the organisation
    @action(detail=True, methods=['get'], url_path='election-stats')
    def election_stats(self, request, pk=None):
        organisation = self.get_object()
        elections = Election.objects.filter(organisation=organisation)
        total_elections = elections.count()
        incoming_elections = elections.filter(date_time_occuring__gt=datetime.now()).count()
        ongoing_election = elections.filter(date_time_occuring__lte=datetime.now(), date_time_ending__gte=datetime.now()).count()
        completed_elections = elections.filter(date_time_ending__lt=datetime.now()).count()


        return Response(
            {
                'total_elections': total_elections,
                'incoming_elections': incoming_elections,
                'ongoing_election': ongoing_election,
                'completed_elections': completed_elections,
            },
            status=status.HTTP_200_OK,
        )

#we will come back to this as membership is what we are looking at exactly
class UserViewset(ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [HasPermission]
    filterset_fields = ['email', 'first_name', 'last_name', 'is_active']

    ACTION_PERMISSION_MAP = {
        'create': 'org.members.manage',
        'update': 'org.members.manage',
        'partial_update': 'org.members.manage',
        'destroy': 'org.members.manage',
    }

    def get_permissions(self):
        if self.action in {'list', 'retrieve'}:
            return [IsAuthenticated()]
        return super().get_permissions()

    #we want to scope to the users in the current membership organisation
    def get_queryset(self):
        user = self.request.user
        
        organisation = get_user_active_organisation(user.id)
        
        return User.objects.filter(memberships__organisation=organisation)
    
    #when we delete we make sure its a soft delete for the user and all their memberships
    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        user.is_active = False
        user.save(update_fields=['is_active'])

        memberships = user.memberships.all()
        for membership in memberships:
            membership.is_active = False
            membership.save(update_fields=['is_active'])
        return Response(status=status.HTTP_204_NO_CONTENT)


class MembershipViewset(ModelViewSet):
    queryset = Membership.objects.all()
    serializer_class = MembershipSerializer
    permission_classes = [HasPermission]
    filterset_fields = ['organisation_id', 'user_id', 'role', 'is_active']

    ACTION_PERMISSION_MAP = {
        'create': 'org.members.manage',
        'bulk_upload': 'org.members.manage',
        'update': 'org.members.manage',
        'partial_update': 'org.members.manage',
        'destroy': 'org.members.manage',
    }

    def get_permissions(self):
        if self.action in {'list', 'retrieve', 'my_memberships', 'switch_membership'}:
            return [IsAuthenticated()]
        return super().get_permissions()
    
    # we want to scope it to the current organisation
    def get_queryset(self):
        user = self.request.user
        if self.action == 'switch_membership':
            return Membership.objects.filter(user=user, is_active=True)
        organisation = get_user_active_organisation(user.id)
        active_membership = get_user_active_membership(user.id)

        #only admin can see all memberships, others can only see their own membership in the organisation
        if active_membership and active_membership.role == 'admin':
            return Membership.objects.filter(organisation=organisation)
        return Membership.objects.filter(organisation=organisation, user=user)
    
    def create(self, request, *args, **kwargs):
        serializer = UserMembershipCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user, created = User.objects.get_or_create(
            email=serializer.validated_data['email'],
            defaults={
                'first_name': serializer.validated_data.get('first_name', ''),
                'last_name': serializer.validated_data.get('last_name', ''),
                'phone': serializer.validated_data.get('phone', ''),
                'bio': serializer.validated_data.get('bio', ''),
            }
        )
        if created:
            user.set_password(serializer.validated_data.get('password', ''))
            user.save(update_fields=['password'])
        membership_is_active = serializer.validated_data.get('is_active', True)
        has_active_membership = Membership.objects.filter(
            user=user,
            currently_active=True,
            is_active=True,
        ).exists()
        membership = Membership.objects.create(
            user=user,
            organisation=serializer.validated_data['organisation'],
            role=serializer.validated_data['role'],
            is_active=membership_is_active,
            currently_active=membership_is_active and not has_active_membership,
        )
        assign_org_default_permissions_to_membership(membership.id, membership.role)
        response_serializer = MembershipSerializer(membership)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        upload = request.FILES.get('file')
        if upload is None:
            return Response({'detail': 'file is required.'}, status=status.HTTP_400_BAD_REQUEST)

        active_organisation = get_user_active_organisation(request.user.id)
        if active_organisation is None:
            return Response({'detail': 'No active organisation found.'}, status=status.HTTP_400_BAD_REQUEST)

        organisation_id = request.data.get('organisation_id')
        if organisation_id not in (None, '', str(active_organisation.id), active_organisation.id):
            return Response(
                {'detail': 'organisation_id must match your active organisation.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            rows = _parse_upload_rows(upload)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        if not rows:
            return Response({'detail': 'The uploaded file has no member rows.'}, status=status.HTTP_400_BAD_REQUEST)

        created_users = 0
        created_memberships = 0
        existing_memberships = 0
        updated_memberships = 0
        skipped_rows = []

        with transaction.atomic():
            for row_number, row in enumerate(rows, start=2):
                email = str(row.get('email') or '').strip().lower()
                if not email:
                    skipped_rows.append({'row': row_number, 'reason': 'email is required'})
                    continue

                role = str(row.get('role') or 'member').strip().lower()
                if role not in ORG_ALLOWED_ROLE_CHOICES:
                    skipped_rows.append({'row': row_number, 'reason': f'invalid role: {role}'})
                    continue

                first_name = str(row.get('first_name') or '').strip()
                last_name = str(row.get('last_name') or '').strip()
                phone = str(row.get('phone') or '').strip() or None
                bio = str(row.get('bio') or '').strip() or None
                is_active = _parse_bool(row.get('is_active'), default=True)

                user = User.objects.filter(email=email).first()
                if user is None:
                    user = User.objects.create_user(
                        email=email,
                        password=get_random_string(32),
                        first_name=first_name,
                        last_name=last_name,
                        phone=phone,
                        bio=bio,
                        is_verified=False,
                    )
                    created_users += 1
                else:
                    user_updates = []
                    if first_name and user.first_name != first_name:
                        user.first_name = first_name
                        user_updates.append('first_name')
                    if last_name and user.last_name != last_name:
                        user.last_name = last_name
                        user_updates.append('last_name')
                    if phone and user.phone != phone:
                        user.phone = phone
                        user_updates.append('phone')
                    if bio and user.bio != bio:
                        user.bio = bio
                        user_updates.append('bio')
                    if user_updates:
                        user.save(update_fields=user_updates)

                has_active_membership = Membership.objects.filter(
                    user=user,
                    currently_active=True,
                    is_active=True,
                ).exists()
                membership, membership_created = Membership.objects.get_or_create(
                    organisation=active_organisation,
                    user=user,
                    defaults={
                        'role': role,
                        'is_active': is_active,
                        'currently_active': is_active and not has_active_membership,
                    },
                )

                if membership_created:
                    created_memberships += 1
                else:
                    existing_memberships += 1
                    updates = []
                    #lets avoid role updates for now
                    # if membership.role != role:
                    #     membership.role = role
                    #     updates.append('role')
                    if membership.is_active != is_active:
                        membership.is_active = is_active
                        updates.append('is_active')
                    has_other_active_membership = Membership.objects.filter(
                        user=user,
                        currently_active=True,
                        is_active=True,
                    ).exclude(id=membership.id).exists()
                    should_be_currently_active = membership.is_active and not has_other_active_membership
                    if membership.currently_active != should_be_currently_active:
                        membership.currently_active = should_be_currently_active
                        updates.append('currently_active')
                    if updates:
                        membership.save(update_fields=updates)
                        updated_memberships += 1

                assign_org_default_permissions_to_membership(membership.id, membership.role)

        response_status = status.HTTP_201_CREATED if created_memberships else status.HTTP_200_OK
        return Response(
            {
                'organisation_id': active_organisation.id,
                'created_users': created_users,
                'created_memberships': created_memberships,
                'existing_memberships': existing_memberships,
                'updated_memberships': updated_memberships,
                'skipped_rows': skipped_rows,
            },
            status=response_status,
        )
    
    def destroy(self, request, *args, **kwargs):
        membership = self.get_object()
        membership.is_active = False
        membership.save(update_fields=['is_active'])
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], url_path='my-memberships')
    def my_memberships(self, request):
        memberships = Membership.objects.filter(user=request.user)
        serializer = self.get_serializer(memberships, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], url_path='switch-membership')
    def switch_membership(self, request, pk=None):
        membership = self.get_object()
        switched_membership = switch_active_membership(request.user.id, membership.id)
        if switched_membership:
            serializer = self.get_serializer(switched_membership)
            return Response(serializer.data)
        return Response({'detail': 'Membership not found or does not belong to user.'}, status=status.HTTP_400_BAD_REQUEST)
    
class PermissionRecordViewset(ModelViewSet):
    queryset = PermissionRecord.objects.all()
    serializer_class = PermissionRecordSerializer
    permission_classes = [HasPermission]


    ACTION_PERMISSION_MAP = {
        'bulk_assign': 'org.access.manage',
        'bulk_unassign': 'org.access.manage',
        'get_membership_permissions': 'org.access.manage',
    }
    
    @action(detail=False, methods=['post'], url_path='bulk-assign')
    def bulk_assign(self, request):
        type = request.data.get('type')
        membership_id = request.data.get('membership_id')
        permissions = request.data.get('permissions')
        election_id = request.data.get('election_id')

        if type == "organisation":
            assign_org_bulk_permissions_to_membership(membership_id, permissions)
        elif type == "election":
            assign_election_bulk_permissions_to_membership(membership_id, permissions, election_id)
        else:            
            return Response({'error': 'Invalid type'}, status=400)
        return Response({'detail': 'Permissions assigned.'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='bulk-unassign')   
    def bulk_unassign(self, request):
        type = request.data.get('type')
        membership_id = request.data.get('membership_id')
        permissions = request.data.get('permissions')
        election_id = request.data.get('election_id')

        if type == "organisation":
            revoke_org_bulk_permissions_from_membership(membership_id, permissions)
        elif type == "election":
            revoke_election_bulk_permissions_from_membership(membership_id, election_id, permissions)
        else:            
            return Response({'error': 'Invalid type'}, status=400)
        return Response({'detail': 'Permissions removed.'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='get-membership-permissions')
    def get_membership_permissions(self, request):
        membership_id = request.query_params.get('membership_id')
        if not membership_id:
            return Response({'detail': 'membership_id is required.'}, status=status.HTTP_400_BAD_REQUEST)
        permissions = get_all_permissions_for_membership(membership_id)
        serializer = self.get_serializer(permissions, many=True)
        return Response(serializer.data)

class ElectionViewset(ModelViewSet):
    queryset = Election.objects.all()
    serializer_class = ElectionSerializer
    permission_classes = [HasPermission]
    filterset_fields = ['organisation_id', 'date_time_occuring', 'date_time_ending', 'winner_id']

    ACTION_PERMISSION_MAP = {
        'create': 'org.elections.manage',
        'update': 'org.elections.manage',
        'partial_update': 'org.elections.manage',
        'destroy': 'org.elections.manage',
        'deploy_contract': 'org.elections.manage',
        'voter_turnout': 'org.analytics.view',
        'send_voter_invites': 'election.invites.manage',
        'enroll_all_members': 'election.participants.manage',
    }

    def get_permissions(self):
        if self.action in {'list', 'retrieve', 'positions', 'participants', 'candidates'}:
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_queryset(self):
        #we want to return elections that a member is a participant or caidate
        user = self.request.user
        membership = get_user_active_membership(user.id)
        if membership and membership.role == 'admin':
            return Election.objects.filter(organisation__memberships=membership).distinct()
        return Election.objects.filter(participants__membership=membership).distinct() | Election.objects.filter(candidates__membership=membership).distinct()
    
    def perform_create(self, serializer):
        user = self.request.user
        membership = get_user_active_membership(user.id)
        election = serializer.save()
        if membership is not None:
            assign_election_default_permissions_to_membership(
                membership.id,
                election.id,
                membership.role,
            )

    #we get the voter turnout in all the elctions of an organisation
    @action(detail=False, methods=['get'], url_path='voter-turnout')
    def voter_turnout(self, request):
        organisation = get_user_active_organisation(request.user.id)
        if organisation is None:
            return Response([], status=status.HTTP_200_OK)

        elections = organisation.elections.all()
        turnout_data = []
        for election in elections:
            total_participants = election.participants.count()
            total_votes = election.votes.count()
            turnout = (total_votes / total_participants * 100) if total_participants > 0 else 0
            turnout_data.append(
                {
                    'id': election.id,
                    'name': election.name,
                    'total_participants': total_participants,
                    'total_votes': total_votes,
                    'turnout': turnout,
                }
            )

        return Response(turnout_data, status=status.HTTP_200_OK)

    #we get the positions of an election
    @action(detail=True, methods=['get'], url_path='election-positions')
    def positions(self, request, pk=None):
        election = self.get_object()
        positions = election.positions.all()
        serializer = PositionSerializer(positions, many=True)
        return Response(serializer.data)
    
    #get all the participants of an election
    @action(detail=True, methods=['get'], url_path='election-participants')
    def participants(self, request, pk=None):
        election = self.get_object()
        participants = election.participants.all()
        serializer = ParticipantSerializer(participants, many=True)
        return Response(serializer.data)
    
    #gets all the candidates of an election
    @action(detail=True, methods=['get'], url_path='election-candidates')
    def candidates(self, request, pk=None):
        election = self.get_object()
        candidates = election.candidates.all()
        serializer = CandidateSerializer(candidates, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='deploy-contract')
    def deploy_contract(self, request, pk=None):
        election = self.get_object()
        already_deployed = bool(election.smart_contract_address)
        if not already_deployed:
            election.smart_contract_address = deploy_election_contract(election)
            election.save(update_fields=['smart_contract_address'])
        return Response(
            {
                'election_id': election.id,
                'smart_contract_address': election.smart_contract_address,
                'already_deployed': already_deployed,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=['post'], url_path='enroll-all-members')
    def enroll_all_members(self, request, pk=None):
        election = self.get_object()
        roles = request.data.get('roles')
        allowed_roles = {role for role, _ in ROLE_CHOICES}

        memberships = Membership.objects.filter(
            organisation=election.organisation,
            is_active=True,
        )

        if roles is not None:
            if not isinstance(roles, list):
                return Response({'detail': 'roles must be a list.'}, status=status.HTTP_400_BAD_REQUEST)
            invalid_roles = [role for role in roles if role not in allowed_roles]
            if invalid_roles:
                return Response(
                    {'detail': f'Invalid roles: {invalid_roles}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            memberships = memberships.filter(role__in=roles)

        created_participants = 0
        existing_participants = 0

        with transaction.atomic():
            for membership in memberships:
                _, created = Participant.objects.get_or_create(
                    election=election,
                    membership=membership,
                )
                if created:
                    created_participants += 1
                else:
                    existing_participants += 1
                assign_election_default_permissions_to_membership(
                    membership.id,
                    election.id,
                    membership.role,
                )

        return Response(
            {
                'election_id': election.id,
                'memberships_processed': memberships.count(),
                'created_participants': created_participants,
                'existing_participants': existing_participants,
                'note': 'Candidates are included by default as long as they are active organisation members.',
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=['post'], url_path='send-voter-invites')
    def send_voter_invites(self, request, pk=None):
        election = self.get_object()
        generated_by = get_user_active_membership(request.user.id)
        summary = dispatch_voter_invites_for_election(
            election=election,
            generated_by=generated_by,
            skip_if_already_dispatched=False,
        )
        return Response(summary, status=status.HTTP_200_OK)


class LogViewset(ModelViewSet):
    queryset = Log.objects.all()
    serializer_class = LogSerializer
    permission_classes = [HasPermission]

    ACTION_PERMISSION_MAP = {
        'membership_logs': 'org.analytics.view',
        'membership_logs_by_election': 'org.analytics.view',
    }


    @action(detail=False, methods=['get'], url_path='membership-logs')
    def membership_logs(self, request):
        logs = Log.objects.filter(membership__user=request.user)
        serializer = self.get_serializer(logs, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='election-logs/(?P<election_id>[^/.]+)')
    def membership_logs_by_election(self, request, election_id=None):
        logs = Log.objects.filter(election_id=election_id, membership__user=request.user)
        serializer = self.get_serializer(logs, many=True)
        return Response(serializer.data)
    
