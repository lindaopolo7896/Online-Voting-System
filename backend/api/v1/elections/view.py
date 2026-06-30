import csv
import io
from pathlib import Path

from django.db import transaction
from django.utils.crypto import get_random_string
from rest_framework.permissions import IsAuthenticated
from rest_framework.viewsets import ModelViewSet
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.elections.models import Position, Participant, Candidate
from api.v1.users.permissions import HasPermission
from apps.users.models import Election, Membership, User
from services.membership_service import get_user_active_membership, get_user_active_organisation
from services.permission_service import assign_election_default_permissions_to_membership
from services.voting_link_service import dispatch_voter_invites_for_election
from .serializers import PositionSerializer, ParticipantSerializer, CandidateSerializer

ALLOWED_ROLE_CHOICES = {'admin', 'official', 'member'}


def _normalize_column_name(value):
    return str(value or '').strip().lower().replace(' ', '_')


def _parse_csv_rows(uploaded_file):
    uploaded_file.seek(0)
    text = uploaded_file.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        if not row or not any(v not in (None, '') for v in row.values()):
            continue
        normalized = {_normalize_column_name(k): v for k, v in row.items()}
        rows.append(normalized)
    return rows


def _parse_xlsx_rows(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('XLSX upload requires openpyxl to be installed.') from exc

    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    raw_rows = list(worksheet.iter_rows(values_only=True))
    if not raw_rows:
        return []

    headers = [_normalize_column_name(header) for header in raw_rows[0]]
    rows = []
    for row in raw_rows[1:]:
        if not row or not any(v not in (None, '') for v in row):
            continue
        normalized = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            normalized[header] = row[idx] if idx < len(row) else None
        rows.append(normalized)
    return rows


def _parse_participant_rows(uploaded_file):
    extension = Path(uploaded_file.name).suffix.lower()
    if extension == '.csv':
        return _parse_csv_rows(uploaded_file)
    if extension in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
        return _parse_xlsx_rows(uploaded_file)
    raise ValueError('Unsupported file type. Please upload .csv or .xlsx.')


class PositionViewSet(ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    permission_classes = [HasPermission]
    filterset_fields = ['organisation_id', 'election_id', 'status', 'name']

    ACTION_PERMISSION_MAP = {
        'create': 'election.ballot.manage',
        'update': 'election.ballot.manage',
        'partial_update': 'election.ballot.manage',
        'destroy': 'election.ballot.manage',
    }

    def get_permissions(self):
        if self.action in {'list', 'retrieve'}:
            return [IsAuthenticated()]
        return super().get_permissions()

    #scope to current election
    def get_queryset(self):
        election_id = self.kwargs.get('election_id')
        organisation = get_user_active_organisation(self.request.user.id)
        if organisation is None:
            return Position.objects.none()
        return Position.objects.filter(election_id=election_id, election__organisation=organisation)

    
class ParticipantViewSet(ModelViewSet):
    queryset = Participant.objects.all()
    serializer_class = ParticipantSerializer
    permission_classes = [HasPermission]
    filterset_fields = ['membership_id', 'election_id', 'has_voted']

    ACTION_PERMISSION_MAP = {
        'create': 'election.participants.manage',
        'bulk_upload': 'election.participants.manage',
        'update': 'election.participants.manage',
        'partial_update': 'election.participants.manage',
        'destroy': 'election.participants.manage',
        'convert_to_candidate': 'election.ballot.manage',
        'send_invitations': 'election.invites.manage',
    }

    def get_permissions(self):
        if self.action in {'list', 'retrieve'}:
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_queryset(self):
        election_id = self.kwargs.get('election_id')
        organisation = get_user_active_organisation(self.request.user.id)
        if organisation is None:
            return Participant.objects.none()
        return Participant.objects.filter(election_id=election_id, election__organisation=organisation)

    def create(self, request, *args, **kwargs):
        election_id = self.kwargs.get('election_id')
        if request.data is None:
            payload = {}
        elif hasattr(request.data, 'copy'):
            payload = request.data.copy()
        elif isinstance(request.data, dict):
            payload = dict(request.data)
        else:
            return Response(
                {'detail': 'Request body must be a JSON object.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        payload['election_id'] = election_id
        serializer = self.get_serializer(data=payload)
        serializer.is_valid(raise_exception=True)
        participant = serializer.save()
        assign_election_default_permissions_to_membership(
            participant.membership_id,
            participant.election_id,
            participant.membership.role,
        )
        response_serializer = self.get_serializer(participant)
        headers = self.get_success_headers(response_serializer.data)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request, election_id=None):
        upload = request.FILES.get('file')
        if upload is None:
            return Response({'detail': 'file is required.'}, status=status.HTTP_400_BAD_REQUEST)

        election = Election.objects.filter(id=election_id).first()
        if election is None:
            return Response({'detail': 'Election not found.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            rows = _parse_participant_rows(upload)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        if not rows:
            return Response({'detail': 'The uploaded file has no participant rows.'}, status=status.HTTP_400_BAD_REQUEST)

        created_users = 0
        created_memberships = 0
        created_participants = 0
        existing_participants = 0
        skipped_rows = []

        with transaction.atomic():
            for row_number, row in enumerate(rows, start=2):
                email = str(row.get('email') or '').strip().lower()
                if not email:
                    skipped_rows.append({'row': row_number, 'reason': 'email is required'})
                    continue

                role = str(row.get('role') or 'member').strip().lower()
                if role not in ALLOWED_ROLE_CHOICES:
                    skipped_rows.append({'row': row_number, 'reason': f'invalid role: {role}'})
                    continue

                first_name = str(row.get('first_name') or '').strip()
                last_name = str(row.get('last_name') or '').strip()
                phone = str(row.get('phone') or '').strip() or None
                bio = str(row.get('bio') or '').strip() or None

                user = User.objects.filter(email=email).first()
                if user is None:
                    temp_password = get_random_string(32)
                    user = User.objects.create_user(
                        email=email,
                        password=temp_password,
                        first_name=first_name,
                        last_name=last_name,
                        phone=phone,
                        bio=bio,
                    )
                    created_users += 1
                has_active_membership = Membership.objects.filter(
                    user=user,
                    currently_active=True,
                    is_active=True,
                ).exists()

                membership, membership_created = Membership.objects.get_or_create(
                    organisation=election.organisation,
                    user=user,
                    defaults={
                        'role': role,
                        'currently_active': not has_active_membership,
                        'is_active': True,
                    },
                )
                if membership_created:
                    created_memberships += 1
                else:
                    updates = []
                    #lets avoid role updates for now
                    # if membership.role != role:
                    #     membership.role = role
                    #     updates.append('role')
                    if not membership.is_active:
                        membership.is_active = True
                        updates.append('is_active')
                    has_other_active_membership = Membership.objects.filter(
                        user=user,
                        currently_active=True,
                        is_active=True,
                    ).exclude(id=membership.id).exists()
                    if membership.is_active and not membership.currently_active and not has_other_active_membership:
                        membership.currently_active = True
                        updates.append('currently_active')
                    if updates:
                        membership.save(update_fields=updates)

                _, participant_created = Participant.objects.get_or_create(
                    election=election,
                    membership=membership,
                )
                if participant_created:
                    created_participants += 1
                else:
                    existing_participants += 1
                assign_election_default_permissions_to_membership(
                    membership.id,
                    election.id,
                    membership.role,
                )

        response_status = status.HTTP_201_CREATED if created_participants else status.HTTP_200_OK
        return Response(
            {
                'created_users': created_users,
                'created_memberships': created_memberships,
                'created_participants': created_participants,
                'existing_participants': existing_participants,
                'skipped_rows': skipped_rows,
            },
            status=response_status,
        )

    @action(detail=False, methods=['post'], url_path='send-invitations')
    def send_invitations(self, request, election_id=None):
        election = Election.objects.filter(id=election_id).first()
        if election is None:
            return Response({'detail': 'Election not found.'}, status=status.HTTP_404_NOT_FOUND)

        generated_by = get_user_active_membership(request.user.id)
        summary = dispatch_voter_invites_for_election(
            election=election,
            generated_by=generated_by,
            skip_if_already_dispatched=False,
        )
        return Response(summary, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='convert-to-candidate')
    def convert_to_candidate(self, request, election_id=None, pk=None):
        participant = self.get_object()
        position_id = request.data.get('position_id')
        if not position_id:
            return Response({'detail': 'position_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

        position = Position.objects.filter(id=position_id, election_id=participant.election_id).first()
        if position is None:
            return Response(
                {'detail': 'Position not found in this election.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if Candidate.objects.filter(
            election_id=participant.election_id,
            membership_id=participant.membership_id,
            position_id=position.id,
        ).exists():
            return Response(
                {'detail': 'Participant is already a candidate for this position.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        candidate_data = {
            'membership_id': participant.membership_id,
            'election_id': participant.election_id,
            'position_id': position.id,
            'manifesto': request.data.get('manifesto') or None,
            'slogan': request.data.get('slogan') or None,
            'status': request.data.get('status') or 'active',
        }
        campaign_photos = request.data.get('campaign_photos')
        if campaign_photos not in (None, ''):
            candidate_data['campaign_photos'] = campaign_photos

        serializer = CandidateSerializer(data=candidate_data)
        serializer.is_valid(raise_exception=True)
        candidate = serializer.save()
        return Response(CandidateSerializer(candidate).data, status=status.HTTP_201_CREATED)


class CandidateViewSet(ModelViewSet):
    queryset = Candidate.objects.all()
    serializer_class = CandidateSerializer
    permission_classes = [HasPermission]
    filterset_fields = ['membership_id', 'election_id', 'position_id', 'status']

    ACTION_PERMISSION_MAP = {
        'create': 'election.ballot.manage',
        'update': 'election.ballot.manage',
        'partial_update': 'election.ballot.manage',
        'destroy': 'election.ballot.manage',
    }

    def get_permissions(self):
        if self.action in {'list', 'retrieve'}:
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_queryset(self):
        election_id = self.kwargs.get('election_id')
        organisation = get_user_active_organisation(self.request.user.id)
        if organisation is None:
            return Candidate.objects.none()
        return Candidate.objects.filter(election_id=election_id, election__organisation=organisation)
